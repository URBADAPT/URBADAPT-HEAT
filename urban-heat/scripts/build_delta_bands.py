"""
Build climate-uncertainty delta CSV from raw PROVIDE xlsx files.

The PROVIDE percentiles (pct05–pct95) represent percentiles of the base
temperature distribution within each month — NOT uncertainty in warming.
- p50 delta = how much the median temperature shifts
- p95 delta = how much the hottest temperatures shift

Therefore uncertainty bands (low/central/high) must come from the spread
ACROSS GCMs at a fixed within-distribution percentile, not from varying
the percentile within a single GCM.

For each variable, a target within-distribution percentile is chosen:
  - tas/tasmin/hurs/sfcWind/ts (and their _std): p50 = avg(pct45, pct55)
  - tasmax (and tasmax_std):                      p80 = avg(pct75, pct85)
    (daily Tmax lives in the upper tail of the distribution)

Then across-GCM uncertainty bands are:
  low     = 25th percentile of per-GCM deltas
  central = median (50th percentile) of per-GCM deltas
  high    = 75th percentile of per-GCM deltas

Output columns (same schema as before):
    city, year, clim_scen, var, month, pct_band, delta

Usage:
    python scripts/build_delta_bands.py
    python scripts/build_delta_bands.py --gcm-low-pct 10 --gcm-high-pct 90

Requires: openpyxl, pandas, numpy
"""

import argparse
import glob
import os
import re
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

# Within-distribution percentile targets per variable 
# For each PROVIDE variable, which pair of percentiles to average as
# the "representative delta" for that variable.
# tas/tasmin → p50 (median temperature shift)
# tasmax → p80 (upper-tail, where Tmax lives)
VAR_PCT_TARGETS = {
    "tas":        ("pct45", "pct55"),
    "tas_std":    ("pct45", "pct55"),
    "tasmin":     ("pct45", "pct55"),
    "tasmin_std": ("pct45", "pct55"),
    "tasmax":     ("pct75", "pct85"),
    "tasmax_std": ("pct75", "pct85"),
    "hurs":       ("pct45", "pct55"),
    "hurs_std":   ("pct45", "pct55"),
    "sfcWind":    ("pct45", "pct55"),
    "sfcWind_std":("pct45", "pct55"),
    "ts":         ("pct45", "pct55"),
    "ts_std":     ("pct45", "pct55"),
}

# Default paths 
DEFAULT_DELTAS_ROOT = Path(__file__).resolve().parent.parent.parent / "future_deltas"
CITIES = ["Rome", "Athens", "Lisbon"]
SCENARIOS = ["CurPol", "GS", "SP", "ssp585"]

VARIABLES = list(VAR_PCT_TARGETS.keys())

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def parse_filename(path: str):
    """Extract city, year, gcm from filename.

    Supports two conventions:
      - Per-GCM:       Rome_2050_CMCC-CM2-SR5_ensmean.xlsx
      - Multi-model:   Rome_2050_ssp585_mean.xlsx  (already averaged across GCMs)
    """
    fname = os.path.basename(path)
    m = re.match(r"(.+?)_(\d{4})_(.+?)_ensmean\.xlsx$", fname)
    if m:
        return {"city": m.group(1), "year": int(m.group(2)), "gcm": m.group(3)}
    m = re.match(r"(.+?)_(\d{4})_(.+?)_mean\.xlsx$", fname)
    if m:
        return {"city": m.group(1), "year": int(m.group(2)), "gcm": "__multimodel__"}
    return None


def read_xlsx_gcm_deltas(path: str, variables: list[str]) -> list[dict]:
    """Read one GCM xlsx and return one delta per variable × month.

    For each variable, extracts the delta at the target within-distribution
    percentile (defined in VAR_PCT_TARGETS).
    """
    meta = parse_filename(path)
    if meta is None:
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = []

    for var in variables:
        if var not in wb.sheetnames:
            continue
        ws = wb[var]
        pct_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            pct_label = str(row[0]).strip()
            values = list(row[1:13])
            pct_data[pct_label] = values

        p1, p2 = VAR_PCT_TARGETS[var]
        if p1 not in pct_data or p2 not in pct_data:
            continue
        v1 = pct_data[p1]
        v2 = pct_data[p2]
        for m_idx in range(12):
            val1 = v1[m_idx]
            val2 = v2[m_idx]
            if val1 is None or val2 is None:
                continue
            avg_val = (float(val1) + float(val2)) / 2.0
            rows.append({
                "city": meta["city"],
                "year": meta["year"],
                "gcm": meta["gcm"],
                "var": var,
                "month": m_idx + 1,
                "delta": avg_val,
            })
    wb.close()
    return rows


def compute_gcm_bands(df: pd.DataFrame, low_pct: float, high_pct: float) -> pd.DataFrame:
    """From per-GCM deltas, compute low/central/high bands using across-GCM percentiles."""
    group_cols = ["city", "year", "clim_scen", "var", "month"]

    df_low = df.groupby(group_cols, as_index=False)["delta"].quantile(low_pct / 100)
    df_low["pct_band"] = "low"

    df_central = df.groupby(group_cols, as_index=False)["delta"].quantile(0.50)
    df_central["pct_band"] = "central"

    df_high = df.groupby(group_cols, as_index=False)["delta"].quantile(high_pct / 100)
    df_high["pct_band"] = "high"

    return pd.concat([df_low, df_central, df_high], ignore_index=True)


def main():
    parser = argparse.ArgumentParser(description="Build climate-uncertainty delta CSV (across-GCM bands)")
    parser.add_argument("--deltas-root", type=Path, default=DEFAULT_DELTAS_ROOT,
                        help="Root folder containing city subfolders with xlsx files")
    parser.add_argument("--cities", nargs="+", default=CITIES)
    parser.add_argument("--scenarios", nargs="+", default=SCENARIOS)
    parser.add_argument("--gcm-low-pct", type=float, default=25,
                        help="Percentile for 'low' band across GCMs (default: 25)")
    parser.add_argument("--gcm-high-pct", type=float, default=75,
                        help="Percentile for 'high' band across GCMs (default: 75)")
    parser.add_argument("--single-output", type=Path, default=None,
                        help="Single combined CSV for all cities")
    args = parser.parse_args()

    print(f"GCM uncertainty bands: low=p{args.gcm_low_pct:.0f}, central=p50, high=p{args.gcm_high_pct:.0f}")
    print(f"Within-distribution percentile targets:")
    seen = set()
    for var, (p1, p2) in VAR_PCT_TARGETS.items():
        label = f"  avg({p1},{p2})"
        if label not in seen:
            vars_with_same = [v for v, (a, b) in VAR_PCT_TARGETS.items() if (a, b) == (p1, p2)]
            print(f"  {', '.join(vars_with_same)}: avg({p1}, {p2})")
            seen.add(label)
    print()

    all_bands = []

    for city in args.cities:
        city_rows = []
        for scen in args.scenarios:
            scen_dir = args.deltas_root / city / scen
            if not scen_dir.exists():
                print(f"  SKIP {scen_dir} (not found)")
                continue

            xlsx_files = sorted(glob.glob(str(scen_dir / f"{city}_*_ensmean.xlsx")))
            xlsx_mm = sorted(glob.glob(str(scen_dir / f"{city}_*_mean.xlsx")))
            xlsx_mm = [f for f in xlsx_mm if "_ensmean.xlsx" not in f]

            # Only use per-GCM files for across-GCM bands
            # Multi-model mean files (ssp585) are already averaged — use as-is
            if xlsx_files:
                n_gcm = len(xlsx_files)
                print(f"  {city}/{scen}: {n_gcm} GCM files")
                for fpath in xlsx_files:
                    rows = read_xlsx_gcm_deltas(fpath, VARIABLES)
                    for r in rows:
                        r["clim_scen"] = scen
                    city_rows.extend(rows)
            elif xlsx_mm:
                # ssp585-style: single multi-model mean file, no GCM spread available
                print(f"  {city}/{scen}: {len(xlsx_mm)} multi-model files (no GCM spread)")
                for fpath in xlsx_mm:
                    rows = read_xlsx_gcm_deltas(fpath, VARIABLES)
                    for r in rows:
                        r["clim_scen"] = scen
                    city_rows.extend(rows)
            else:
                print(f"  SKIP {scen_dir} (no xlsx files)")

        if not city_rows:
            print(f"  WARNING: no data for {city}")
            continue

        df = pd.DataFrame(city_rows)

        # Separate per-GCM data from multi-model data
        df_gcm = df[df["gcm"] != "__multimodel__"].copy()
        df_mm = df[df["gcm"] == "__multimodel__"].copy()

        frames = []

        if not df_gcm.empty:
            n_gcms = df_gcm["gcm"].nunique()
            bands_gcm = compute_gcm_bands(df_gcm, args.gcm_low_pct, args.gcm_high_pct)
            bands_gcm["n_gcms"] = n_gcms
            frames.append(bands_gcm)
            print(f"  {city}: {len(bands_gcm)} rows from {n_gcms} GCMs")

        if not df_mm.empty:
            # Multi-model files: only central band (no spread to compute)
            df_mm_out = df_mm[["city", "year", "clim_scen", "var", "month", "delta"]].copy()
            df_mm_out["pct_band"] = "central"
            df_mm_out["n_gcms"] = 1
            frames.append(df_mm_out)
            print(f"  {city}: {len(df_mm_out)} rows from multi-model mean (central only)")

        if not frames:
            continue

        df_bands = pd.concat(frames, ignore_index=True)
        all_bands.append(df_bands)

        # Write per-city CSV
        urban_heat = Path(__file__).resolve().parent.parent
        city_dir = urban_heat / "data" / city / "T2MmeanDeltas"
        city_dir.mkdir(parents=True, exist_ok=True)
        out_path = city_dir / "climate_change_provide_markups_bands.csv"
        df_bands.drop(columns=["n_gcms"]).to_csv(out_path, index=False)
        print(f"  -> {out_path}")

    # write single combined CSV
    if args.single_output and all_bands:
        df_all = pd.concat(all_bands, ignore_index=True)
        df_all.drop(columns=["n_gcms"]).to_csv(args.single_output, index=False)
        print(f"\nCombined CSV: {args.single_output}")

    # Diagnostic: show JJA spread for tas to verify
    print("\n── Diagnostic: JJA tas deltas (central band, CurPol, 2050) ──")
    for city in args.cities:
        urban_heat = Path(__file__).resolve().parent.parent
        bands_path = urban_heat / "data" / city / "T2MmeanDeltas" / "climate_change_provide_markups_bands.csv"
        if not bands_path.exists():
            continue
        df_b = pd.read_csv(bands_path)
        sub = df_b[
            (df_b["var"] == "tas")
            & (df_b["clim_scen"] == "CurPol")
            & (df_b["year"] == 2050)
            & (df_b["month"].isin([6, 7, 8]))
        ]
        if sub.empty:
            print(f"  {city}: no data")
            continue
        piv = sub.pivot_table(index="month", columns="pct_band", values="delta")
        month_map = {6: "Jun", 7: "Jul", 8: "Aug"}
        piv.index = piv.index.map(month_map)
        print(f"  {city}:")
        for col in ["low", "central", "high"]:
            if col in piv.columns:
                vals = piv[col]
                print(f"    {col:>7}: {vals.values}")


if __name__ == "__main__":
    main()
